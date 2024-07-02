from copiar import copiar_para_destino
from apagar import apagar_colunas
from openpyxl import load_workbook
import shutil

file_path = r"W:/Trigg/Acionamentos Massa/Ocorrencias Whatsapp/ok/LAYOUT AÇÃO DE WHATS TRIGG.xlsx"

apagar_colunas(file_path)

# Caminho do arquivo Excel de destino (LAYOUT AÇÃO DE WHATS TRIGG.xlsx)
dest_file_path = r"W:/Trigg/Acionamentos Massa/Ocorrencias Whatsapp/ok/LAYOUT AÇÃO DE WHATS TRIGG.xlsx"

# Carregar o arquivo de destino
dest_wb = load_workbook(dest_file_path)
dest_sheet = dest_wb.active  # ou selecione a planilha desejada, dependendo da necessidade

# Lista de colunas de origem e destino
colunas = [
    ("R", "A"),
    ("T", "B"),
    ("V", "C"),
    ("AF", "D"),
    ("M", "E"),
    ("AP", "F")
]

# Copiar dados de cada coluna especificada
for source_col, dest_col in colunas:
    copiar_para_destino(
        r"L:/ROBOS/ROBO OCORRENCIA WHATSAPP/MATRIZ/MATRIZ.xlsx",
        "Analitico Contrato",
        source_col,
        dest_col,
        dest_sheet
    )
    

    
# Mensagem a ser inserida na coluna G
mensagem = "Disparo Whatsapp enviado nesta data para retorno em RECEPTIVO Whatsapp OFICIAL e direcionamento ao Portal."

# Encontrar a última linha preenchida nas colunas A, B, C, D, E, F
max_row = max(dest_sheet.max_row, 2)  # Garante que comece da linha 2 caso não haja dados preexistentes

# Preencher a coluna G até a última linha preenchida
for row_idx in range(2, max_row + 1):
    dest_sheet[f"G{row_idx}"].value = mensagem

# Salvar as alterações no arquivo de destino
dest_wb.save(dest_file_path)

# --------------------------------------------------------
# Mover o arquivo para a pasta W:\Trigg\Acionamentos Massa\Ocorrencias Whatsapp
# --------------------------------------------------------

# Novo caminho de destino
novo_destino = r"W:/Trigg/Acionamentos Massa/Ocorrencias Whatsapp"

# Mover o arquivo
shutil.move(dest_file_path, novo_destino)

print("Dados foram copiados para as colunas A, B, C, D, E, F da planilha de destino na linha 2.")
print(f"O arquivo foi movido para: {novo_destino}")