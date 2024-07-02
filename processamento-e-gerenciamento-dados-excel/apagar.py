from openpyxl import load_workbook

# Caminho do arquivo Excel

def apagar_colunas(arquivo):
    # Carregar o arquivo Excel
    wb = load_workbook(arquivo)

    # Selecionar a planilha desejada (por exemplo, a planilha ativa)
    sheet = wb.active

    # Limpar o conteúdo das colunas A a F da linha 2 em diante
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=6):
        for cell in row:
            cell.value = None

    # Salvar as alterações de volta no arquivo
    wb.save(arquivo)

    print("Conteúdo das colunas A a F da linha 2 em diante foi apagado com sucesso.")
