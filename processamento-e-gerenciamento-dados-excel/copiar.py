from openpyxl import load_workbook
import shutil


# --------------------------------------------------------
# Função para copiar dados de uma coluna da planilha de origem para a planilha de destino
# --------------------------------------------------------

def copiar_para_destino(source_file, source_sheet_name, source_column, dest_column, dest_sheet):
    # Carregar o arquivo de origem
    source_wb = load_workbook(source_file, data_only=True)
    source_sheet = source_wb[source_sheet_name]

    # Linha inicial para copiar os dados (linha 3 em diante)
    source_start_row = 3

    # Valor da célula na coluna de origem
    source_values = [source_sheet[f"{source_column}{row_idx}"].value for row_idx in range(source_start_row, source_sheet.max_row + 1)]

    # Preencher a coluna de destino a partir da linha 2
    for idx, value in enumerate(source_values):
        dest_sheet[f"{dest_column}{idx + 2}"].value = value
